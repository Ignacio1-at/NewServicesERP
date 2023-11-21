from django.utils import timezone
from django.shortcuts import render, redirect, get_object_or_404, HttpResponse
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from .forms import CustomLoginForm, FichaNavioForm, FichaPersonalForm, FichaQuimicoForm, FichaVehiculoForm, FichaHerramientasForm, FichaMantenimientoForm
from .models import FichaNavio, FichaPersonal, FichaQuimico, FichaVehiculo, FichaHerramientas, FichaMantenimiento
from django.http import JsonResponse
from django.urls import reverse
from django.views.decorators.http import require_POST
from django.core.serializers import serialize
from django.core.exceptions import ValidationError
from django.forms.utils import ErrorDict
import logging

logger = logging.getLogger(__name__)

#---Excel
import openpyxl
from django.http import HttpResponse, JsonResponse
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
#---------------------------------------------------

def home(request):
    return render(request, 'html/home.html')

#--------------Login
def login_view(request):
    if request.user.is_authenticated:
        return redirect('erp:menu')

    if request.method == 'POST':
        form = CustomLoginForm(request.POST)
        if form.is_valid():
            email = form.cleaned_data['email']
            password = form.cleaned_data['password']
            user = authenticate(request, username=email, password=password)
            if user is not None:
                login(request, user)
                return redirect('erp:menu')
            else:
                messages.error(request, "Correo o contraseña incorrectos. Por favor, inténtalo de nuevo.")

    form = CustomLoginForm()
    return render(request, 'html/login.html', {'form': form})

@login_required
def logout_view(request):
    logout(request)
    return redirect('erp:login')

#-----------------Menu
@login_required
def menu_view(request):
    nombre_usuario = request.user.nombre if request.user.is_authenticated else "Invitado"
    return render(request, 'html/menu.html', {'nombre_usuario': nombre_usuario})

#-----------------Operaciones
@login_required
def gestorOperaciones(request):
    fichas = FichaNavio.objects.all()
    for ficha in fichas:
        print(f"Ficha: {ficha}, Fecha de Creación: {ficha.fecha_creacion}")
    
    nombre_usuario = request.user.nombre if request.user.is_authenticated else "Invitado"

    if request.method == 'POST':
        form = FichaNavioForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Ficha de Navio guardada exitosamente.')
            return redirect('erp:gestor-operaciones')
        else:
            messages.error(request, 'Error en el formulario. Por favor, verifica los datos.')
    else:
        form = FichaNavioForm()

    return render(request, 'html/gestorOperaciones.html', {'form': form, 'fichas': fichas, 'nombre_usuario': nombre_usuario})

@login_required
def eliminar_ficha(request, ficha_id):
    ficha = get_object_or_404(FichaNavio, id=ficha_id)
    ficha.delete()
    messages.success(request, 'Ficha de Navio eliminada exitosamente.')
    return redirect('erp:gestor-operaciones')

def obtener_color_para_estado(estado):
    colores_por_estado = {
        'Terminado': '#00FF00',  # Verde
        'En Proceso': '#FFFF00',  # Amarillo
        'No Iniciado': '#FF0000',  # Rojo
    }

    # Devuelve el color asociado al estado, o blanco por defecto
    return colores_por_estado.get(estado, '#FF0000')

@require_POST
def actualizar_estado(request, ficha_id):
    print(f'Ficha ID: {ficha_id}')
    try:
        ficha = FichaNavio.objects.get(pk=ficha_id)
        nuevo_estado = request.POST.get('nuevo_estado')

        # Realiza la actualización del estado en la base de datos
        ficha.Estado = nuevo_estado
        ficha.color = obtener_color_para_estado(nuevo_estado)
        ficha.save()

        # Devuelve la ficha actualizada en formato JSON
        ficha_actualizada_json = serialize('json', [ficha])

        return JsonResponse({'mensaje': 'Estado actualizado correctamente', 'ficha': ficha_actualizada_json})
    except FichaNavio.DoesNotExist:
        return JsonResponse({'mensaje': 'La ficha no existe'}, status=404)
    except Exception as e:
        return JsonResponse({'mensaje': str(e)}, status=500)

@login_required
def nueva_ficha(request):
    if request.method == 'POST':
        form = FichaNavioForm(request.POST)

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            # En caso de una solicitud AJAX, no es necesario validar 'Estado' y 'color'
            form.fields['Estado'].required = False
            form.fields['color'].required = False

        if form.is_valid():
            ficha = form.save(commit=False)
            ficha.fecha_creacion = timezone.localtime(timezone.now())

            # Establecer valores predeterminados para 'Estado' y 'color'
            ficha.Estado = 'No Iniciado' 
            ficha.color = obtener_color_para_estado(ficha.Estado)

            ficha.save()

            # Verificar si la solicitud es AJAX
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                # Si es una solicitud AJAX, devolver un objeto JSON con la URL a la que redirigir
                return JsonResponse({
                    'mensaje': 'Ficha de Navio guardada exitosamente.',
                    'redireccionar_a': reverse('erp:gestor-operaciones')
                })
            else:
                # Si no es una solicitud AJAX, enviar mensaje de éxito y redirigir normalmente
                messages.success(request, 'Ficha de Navio guardada exitosamente.')
                return redirect('erp:gestor-operaciones')
        else:
            # Si el formulario no es válido, recoger los errores
            errors = form.errors

            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                # Si es una solicitud AJAX, devolver los errores en formato JSON
                return JsonResponse({'errores_validacion': errors}, status=400)
            else:
                # Si no es una solicitud AJAX, enviar los errores al formulario de nuevo
                messages.error(request, 'Error en el formulario. Por favor, verifica los datos.')
                return render(request, 'html/fichaNavio.html', {'form': form})

    elif request.method == 'GET':
        # Si es una solicitud GET, simplemente mostrar el formulario vacío
        form = FichaNavioForm()
        return render(request, 'html/fichaNavio.html', {'form': form})

    else:
        # Si se recibe un método HTTP diferente a GET o POST, loggear el error y redirigir
        logger.error(f"Método HTTP no permitido: {request.method}")
        return redirect('erp:gestor-operaciones')
    
def descargar_excel(request, ficha_id):
    # Obtener datos de la base de datos
    fichas = FichaNavio.objects.filter(id=ficha_id)

    # Crear un nuevo libro de trabajo y una hoja de cálculo
    wb = openpyxl.Workbook()
    ws = wb.active

    # Agregar encabezados a la primera fila
    headers = ["Nave", "Viaje", "Puerto", "Carga", "Procedencia", "Tipo de Servicio",
               "Armador", "Agencia", "Proximo Puerto", "Encalado", "ETA", "Hora Registro ETA",
               "Bomba Sumergible", "Cubierta", "Shape Box", "PCR", "Hora Registro PCR",
               "Fecha Creacion", "Cantidad de Personas", "Puerto"]
    
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f"{col_letter}1"] = header
        ws[f"{col_letter}1"].font = Font(bold=True)

    # Agregar datos a las filas siguientes
    for row_num, ficha in enumerate(fichas, 2):
        ws[f"A{row_num}"] = ficha.Nave
        ws[f"B{row_num}"] = ficha.Viaje
        ws[f"C{row_num}"] = ficha.Puerto
        ws[f"D{row_num}"] = ficha.Carga
        ws[f"E{row_num}"] = ficha.Procedencia
        ws[f"F{row_num}"] = ficha.TipoServicio
        ws[f"G{row_num}"] = ficha.Armador
        ws[f"H{row_num}"] = ficha.Agencia
        ws[f"I{row_num}"] = ficha.ProximoPuerto
        ws[f"J{row_num}"] = ficha.Encalado
        ws[f"K{row_num}"] = ficha.ETA.strftime('%d-%m') if ficha.ETA else None
        ws[f"L{row_num}"] = ficha.horaRegistroETA.strftime('%H:%M') if ficha.horaRegistroETA else None
        ws[f"M{row_num}"] = ficha.Bombasumergible
        ws[f"N{row_num}"] = ficha.Cubierta
        ws[f"O{row_num}"] = ficha.ShapeBox
        ws[f"P{row_num}"] = ficha.PCR
        ws[f"Q{row_num}"] = ficha.horaRegistroPCR.strftime('%H:%M') if ficha.horaRegistroPCR else None
        ws[f"R{row_num}"] = ficha.fecha_creacion.strftime('%d-%m-%Y %H:%M:%S')
        ws[f"S{row_num}"] = ficha.cantidadPersonas
        ws[f"T{row_num}"] = ficha.CantidadPuerto

    # Crear una respuesta de archivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=ficha_navio_{ficha.id}.xlsx'
    wb.save(response)

    return response


#-------------RRHH

@login_required
def gestorPersonal(request):
    fichas = FichaPersonal.objects.all()
    for ficha in fichas:
        print(f"Ficha: {ficha}")
    
    nombre_usuario = request.user.nombre if request.user.is_authenticated else "Invitado"

    if request.method == 'POST':
        form = FichaPersonalForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Ficha de Personal guardada exitosamente.')
            return redirect('erp:gestor-personal')
        else:
            messages.error(request, 'Error en el formulario. Por favor, verifica los datos.')

    else:
        form = FichaPersonalForm()

    return render(request, 'html/gestorPersonal.html', {'form': form, 'fichas': fichas, 'nombre_usuario': nombre_usuario})

def obtener_color_para_estadoPersonal(estado):
    colores_por_estado = {
        'Disponible': '#00FF00',  # Verde
        'En Operacion': '#FFFF00',  # Amarillo
        'No Disponible': '#FF0000',  # Rojo
    }

    # Devuelve el color asociado al estado, o blanco por defecto
    return colores_por_estadoPersonal.get(estado, '#FF0000')

@login_required
def nueva_fichaPersonal(request):
    errors = {}

    if request.method == 'POST':
        form = FichaPersonalForm(request.POST)

        if form.is_valid():
            ficha = form.save(commit=False)
            ficha.Estado = 'No Iniciado'
            ficha.color = obtener_color_para_estadoPersonal(ficha.Estado)
            form.instance.Estado = ficha.Estado
            form.instance.color = ficha.color
            form.save()

            messages.success(request, 'Ficha de Personal guardada exitosamente.')

            # Devuelve una respuesta JSON con la URL a la que se redirigirá
            return JsonResponse({'redireccionar_a': reverse('erp:gestor-personal')})
        else:
            errors = form.errors.as_json()

    # Devuelve una respuesta JSON con los errores si el formulario no es válido
    return JsonResponse({'mensaje': 'Error en el formulario', 'errores_validacion': errors}, status=400)

#----------Inventario
@login_required
def inventario(request):
    nombre_usuario = request.user.nombre if request.user.is_authenticated else "Invitado"
    return render(request, 'html/MenuInventario.html', {'nombre_usuario': nombre_usuario})

#-------Herramientas
@login_required
def gestorHerramientas(request):
    fichas = FichaHerramientas.objects.all()
    for ficha in fichas:
        print(f"Ficha: {ficha}, Fecha de Ingreso: {ficha.fecha_ingreso}")

    nombre_usuario = request.user.nombre if request.user.is_authenticated else "Invitado"

    if request.method == 'POST':
        form = FichaVehiculoForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Ficha de herramientas guardada exitosamente.')
            return redirect('erp:gestor-herramientas')  # Redirige a la página 'gestor-quimico'
        else:
            messages.error(request, 'Error en el formulario. Por favor, verifica los datos.')
    else:
        form = FichaVehiculoForm()

    return render(request, 'html/gestorHerramientas.html', {'form': form, 'fichas': fichas, 'nombre_usuario': nombre_usuario})

@login_required
def nuevaFichaHerramientas(request):
    if request.method == 'POST':
        form = FichaHerramientasForm(request.POST)
        if form.is_valid():
            ficha = form.save(commit=False)
            ficha.fechaIngreso = timezone.now().date()
            print("Valores del formulario antes de guardar:", form.cleaned_data)

            ficha.save()

            messages.success(request, 'Ficha de herramientas guardada exitosamente.')

            print(f"Ficha guardada: {ficha}")

            # Redirige directamente desde Django
            return redirect('erp:gestor-herramientas')  # Redirige a la página 'gestor-quimico'
        else:
            print("Errores de validación del formulario:", form.errors)
            messages.error(request, 'Error en el formulario. Por favor, verifica los datos.')
    else:
        form = FichaVehiculoForm()

    return render(request, 'html/fichaHerramienta.html', {'form': form})
def descargar_excelHerramientas(request):
    # Obtener datos de la base de datos
    fichas = FichaHerramientas.objects.all()

    # Crear un nuevo libro de trabajo y una hoja de cálculo
    wb = openpyxl.Workbook()
    ws = wb.active

    # Agregar encabezados a la primera fila
    headers = ["marca", "fecha_ingreso", "modelo", "cantidad_herramientas", "tipo_herramienta"]

    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f"{col_letter}1"] = header
        ws[f"{col_letter}1"].font = Font(bold=True)

    # Agregar datos a las filas siguientes
    for row_num, ficha in enumerate(fichas, 2):
        ws[f"A{row_num}"] = ficha.marca
        ws[f"B{row_num}"] = ficha.fecha_ingreso
        ws[f"C{row_num}"] = ficha.modelo
        ws[f"D{row_num}"] = ficha.cantidad_herramientas
        ws[f"E{row_num}"] = ficha.tipo_herramienta

    # Crear una respuesta de archivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=fichas_herramientas.xlsx'
    wb.save(response)

    return response

def eliminar_fichaHerramientas(request, ficha_id):
    ficha = get_object_or_404(FichaHerramientas, id=ficha_id)
    ficha.delete()
    messages.success(request, 'Ficha de herramientas eliminada exitosamente.')
    return redirect('erp:gestor-herramientas')

#-------------Quimicos
def gestorQuimico(request):  
    fichas = FichaQuimico.objects.all()
    for ficha in fichas:
        print(f"Ficha: {ficha}, Fecha de Registro: {ficha.fecha_registro}")

    nombre_usuario = request.user.nombre if request.user.is_authenticated else "Invitado"

    if request.method == 'POST':
        form = FichaQuimicoForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Ficha de Quimicos guardada exitosamente.')
            return redirect('erp:gestor-operaciones')
        else:
            messages.error(request, 'Error en el formulario. Por favor, verifica los datos.')
    else:
        form = FichaQuimicoForm()

    return render(request, 'html/gestorQuimico.html', {'form': form, 'fichas': fichas, 'nombre_usuario': nombre_usuario})

@login_required
def nuevaFichaQuimico(request):
    if request.method == 'POST':
        form = FichaQuimicoForm(request.POST)
        if form.is_valid():
            ficha = form.save(commit=False)
            ficha.fecha_registro = timezone.now().date()
            print("Valores del formulario antes de guardar:", form.cleaned_data)

            ficha.save()

            messages.success(request, 'Ficha de Quimico guardada exitosamente.')

            print(f"Ficha guardada: {ficha}")

            # Redirige directamente desde Django
            return redirect('erp:gestor-quimico')  # Redirige a la página 'gestor-quimico'
        else:
            print("Errores de validación del formulario:", form.errors)
            messages.error(request, 'Error en el formulario. Por favor, verifica los datos.')
    else:
        form = FichaQuimicoForm()

    return render(request, 'html/fichaQuimico.html', {'form': form})

def eliminar_fichaQuimico(request, ficha_id):
    ficha = get_object_or_404(FichaQuimico, id=ficha_id)
    ficha.delete()
    messages.success(request, 'Ficha de quimico eliminada exitosamente.')
    return redirect('erp:gestor-quimico')

def descargar_excelQuimico(request):
    # Obtener datos de la base de datos
    fichas = FichaQuimico.objects.all()

    # Crear un nuevo libro de trabajo y una hoja de cálculo
    wb = openpyxl.Workbook()
    ws = wb.active

    # Agregar encabezados a la primera fila
    headers = ["tipo_quimico", "fecha_registro", "capacidad_bines", "lugar_almacenamiento"]

    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f"{col_letter}1"] = header
        ws[f"{col_letter}1"].font = Font(bold=True)

    # Agregar datos a las filas siguientes
    for row_num, ficha in enumerate(fichas, 2):
        ws[f"A{row_num}"] = ficha.tipo_quimico
        ws[f"B{row_num}"] = ficha.fecha_registro
        ws[f"C{row_num}"] = ficha.capacidad_bines
        ws[f"D{row_num}"] = ficha.lugar_almacenamiento

    # Crear una respuesta de archivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=fichas_quimica.xlsx'
    wb.save(response)

    return response

#-----Vehiculo

def gestorVehiculo(request):
  fichas = FichaVehiculo.objects.all()
  for ficha in fichas:
      print(f"Ficha: {ficha}, Fecha de Ingreso: {ficha.fecha_ingreso}")

  nombre_usuario = request.user.nombre if request.user.is_authenticated else "Invitado"

  if request.method == 'POST':
      form = FichaVehiculoForm(request.POST)
      if form.is_valid():
          form.save()
          messages.success(request, 'Ficha de Vehículo guardada exitosamente.')
          return redirect('erp:gestor-vehiculo')  # Redirige a la página 'gestor-quimico'
      else:
          messages.error(request, 'Error en el formulario. Por favor, verifica los datos.')
  else:
      form = FichaVehiculoForm()

  return render(request, 'html/gestorVehiculo.html', {'form': form, 'fichas': fichas, 'nombre_usuario': nombre_usuario})

@login_required
def nuevaFichaVehiculo(request):
  if request.method == 'POST':
      form = FichaVehiculoForm(request.POST)
      if form.is_valid():
          ficha = form.save(commit=False)
          ficha.fecha_ingreso = timezone.now().date()
          print("Valores del formulario antes de guardar:", form.cleaned_data)

          ficha.save()

          messages.success(request, 'Ficha de Vehiculo guardada exitosamente.')

          print(f"Ficha guardada: {ficha}")

          # Redirige directamente desde Django
          return redirect('erp:gestor-vehiculo')  # Redirige a la página 'gestor-quimico'
      else:
          print("Errores de validación del formulario:", form.errors)
          messages.error(request, 'Error en el formulario. Por favor, verifica los datos.')
  else:
      form = FichaVehiculoForm()

  return render(request, 'html/fichaVehiculos.html', {'form': form})

def eliminar_fichaVehiculo(request, ficha_id):
  ficha = get_object_or_404(FichaVehiculo, id=ficha_id)
  ficha.delete()
  messages.success(request, 'Ficha de vehiculo eliminada exitosamente.')
  return redirect('erp:gestor-vehiculo')

def descargar_excelVehiculo(request):
  # Obtener datos de la base de datos
  fichas = FichaVehiculo.objects.all()

  # Crear un nuevo libro de trabajo y una hoja de cálculo
  wb = openpyxl.Workbook()
  ws = wb.active

  # Agregar encabezados a la primera fila
  headers = ["marca", "modelo", "patente", "fecha_ingreso", "chasis", "tipo_vehiculo", "tipo_combustible"]

  for col_num, header in enumerate(headers, 1):
      col_letter = get_column_letter(col_num)
      ws[f"{col_letter}1"] = header
      ws[f"{col_letter}1"].font = Font(bold=True)

  # Agregar datos a las filas siguientes
  for row_num, ficha in enumerate(fichas, 2):
      ws[f"A{row_num}"] = ficha.marca
      ws[f"B{row_num}"] = ficha.modelo
      ws[f"C{row_num}"] = ficha.patente
      ws[f"D{row_num}"] = ficha.fecha_ingreso
      ws[f"E{row_num}"] = ficha.chasis
      ws[f"F{row_num}"] = ficha.tipo_vehiculo
      ws[f"G{row_num}"] = ficha.tipo_combustible

  # Crear una respuesta de archivo Excel
  response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
  response['Content-Disposition'] = 'attachment; filename=fichas_vehiculos.xlsx'
  wb.save(response)

  return response

#-------Herramientas

@login_required
def gestorHerramientas(request):
    fichas = FichaHerramientas.objects.all()
    for ficha in fichas:
        print(f"Ficha: {ficha}, Fecha de Ingreso: {ficha.fecha_ingreso}")

    nombre_usuario = request.user.nombre if request.user.is_authenticated else "Invitado"

    if request.method == 'POST':
        form = FichaVehiculoForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Ficha de herramientas guardada exitosamente.')
            return redirect('erp:gestor-herramientas')  # Redirige a la página 'gestor-quimico'
        else:
            messages.error(request, 'Error en el formulario. Por favor, verifica los datos.')
    else:
        form = FichaVehiculoForm()

    return render(request, 'html/gestorHerramientas.html', {'form': form, 'fichas': fichas, 'nombre_usuario': nombre_usuario})

@login_required
def nuevaFichaHerramientas(request):
    if request.method == 'POST':
        form = FichaHerramientasForm(request.POST)
        if form.is_valid():
            ficha = form.save(commit=False)
            ficha.fechaIngreso = timezone.now().date()
            print("Valores del formulario antes de guardar:", form.cleaned_data)

            ficha.save()

            messages.success(request, 'Ficha de herramientas guardada exitosamente.')

            print(f"Ficha guardada: {ficha}")

            # Redirige directamente desde Django
            return redirect('erp:gestor-herramientas')  # Redirige a la página 'gestor-quimico'
        else:
            print("Errores de validación del formulario:", form.errors)
            messages.error(request, 'Error en el formulario. Por favor, verifica los datos.')
    else:
        form = FichaVehiculoForm()

    return render(request, 'html/fichaHerramientas.html', {'form': form})
def descargar_excelHerramientas(request):
    # Obtener datos de la base de datos
    fichas = FichaHerramientas.objects.all()

    # Crear un nuevo libro de trabajo y una hoja de cálculo
    wb = openpyxl.Workbook()
    ws = wb.active

    # Agregar encabezados a la primera fila
    headers = ["marca", "fecha_ingreso", "modelo", "cantidad_herramientas", "tipo_herramienta"]

    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f"{col_letter}1"] = header
        ws[f"{col_letter}1"].font = Font(bold=True)

    # Agregar datos a las filas siguientes
    for row_num, ficha in enumerate(fichas, 2):
        ws[f"A{row_num}"] = ficha.marca
        ws[f"B{row_num}"] = ficha.fecha_ingreso
        ws[f"C{row_num}"] = ficha.modelo
        ws[f"D{row_num}"] = ficha.cantidad_herramientas
        ws[f"E{row_num}"] = ficha.tipo_herramienta

    # Crear una respuesta de archivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=fichas_herramientas.xlsx'
    wb.save(response)

    return response

def eliminar_fichaHerramientas(request, ficha_id):
    ficha = get_object_or_404(FichaHerramientas, id=ficha_id)
    ficha.delete()
    messages.success(request, 'Ficha de herramientas eliminada exitosamente.')
    return redirect('erp:gestor-herramientas')


def gestorMantenimiento(request):
  fichas = FichaMantenimiento.objects.all()
  for ficha in fichas:
      print(f"Ficha: {ficha}, Fecha de Ingreso: {ficha.fecha_ingreso}")

  nombre_usuario = request.user.nombre if request.user.is_authenticated else "Invitado"

  if request.method == 'POST':
      form = FichaMantenimientoForm(request.POST)
      if form.is_valid():
          form.save()
          messages.success(request, 'Ficha de Mantenimiento guardada exitosamente.')
          return redirect('erp:mantenimiento')  # Redirige a la página 'Mantenimiento'
      else:
          messages.error(request, 'Error en el formulario. Por favor, verifica los datos.')
  else:
      form = FichaMantenimientoForm()

  return render(request, 'html/Mantenimiento.html', {'form': form, 'fichas': fichas, 'nombre_usuario': nombre_usuario})

#---------------------Mantenimiento

def gestorMantenimiento(request):
  fichas = FichaMantenimiento.objects.all()
  for ficha in fichas:
      print(f"Ficha: {ficha}, Fecha de Ingreso: {ficha.fecha_ingreso}")

  nombre_usuario = request.user.nombre if request.user.is_authenticated else "Invitado"

  if request.method == 'POST':
      form = FichaMantenimientoForm(request.POST)
      if form.is_valid():
          form.save()
          messages.success(request, 'Ficha de Mantenimiento guardada exitosamente.')
          return redirect('erp:GestorMantenimiento')  # Redirige a la página 'Mantenimiento'
      else:
          messages.error(request, 'Error en el formulario. Por favor, verifica los datos.')
  else:
      form = FichaMantenimientoForm()

  return render(request, 'html/GestorMantenimiento.html', {'form': form, 'fichas': fichas, 'nombre_usuario': nombre_usuario})

@login_required
def nuevaFichaMantenimiento(request):
  if request.method == 'POST':
      form = nuevaFichaMantenimientoForm(request.POST)
      if form.is_valid():
          ficha = form.save(commit=False)
          ficha.fecha_ingreso = timezone.now().date()
          print("Valores del formulario antes de guardar:", form.cleaned_data)

          ficha.save()

          messages.success(request, 'Ficha de Mantenimiento guardada exitosamente.')

          print(f"Ficha guardada: {ficha}")

          # Redirige directamente desde Django
          return redirect('erp:nuevaFichaMantenimiento')  # Redirige a la página 'Mantenimiento'
      else:
          print("Errores de validación del formulario:", form.errors)
          messages.error(request, 'Error en el formulario. Por favor, verifica los datos.')
  else:
      form = FichaMantenimientoForm()

  return render(request, 'html/FichaMantenimiento.html', {'form': form})

def eliminar_fichaMantenimiento(request, ficha_id):
  ficha = get_object_or_404(FichaMantenimiento, id=ficha_id)
  ficha.delete()
  messages.success(request, 'Ficha de Mantenimiento eliminada exitosamente.')
  return redirect('erp:nuevaFichaMantenimiento')

def descargar_excelMantenimiento(request):
  # Obtener datos de la base de datos
  fichas = FichaMantenimiento.objects.all()

  # Crear un nuevo libro de trabajo y una hoja de cálculo
  wb = openpyxl.Workbook()
  ws = wb.active

  # Agregar encabezados a la primera fila
  headers = ["categoria", "accion", "estado"]

  for col_num, header in enumerate(headers, 1):
      col_letter = get_column_letter(col_num)
      ws[f"{col_letter}1"] = header
      ws[f"{col_letter}1"].font = Font(bold=True)

  # Agregar datos a las filas siguientes
  for row_num, ficha in enumerate(fichas, 2):
      ws[f"A{row_num}"] = ficha.categoria
      ws[f"B{row_num}"] = ficha.accion
      ws[f"C{row_num}"] = ficha.estado

  # Crear una respuesta de archivo Excel
  response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
  response['Content-Disposition'] = 'attachment; filename=Mantenimiento.xlsx'
  wb.save(response)

  return response

#---------------------Documentos

def gestor_documentos(request):
    return render(request, 'html/gestorDocumentos.html')